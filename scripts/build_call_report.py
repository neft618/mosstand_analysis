from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.trendline import Trendline
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
INPUT_XLSX = ROOT / "input" / "звонки_менеджеры_март-05июня.xlsx"
INPUT_JSON_DIR = ROOT / "input" / "calls_json"
OUTPUT_DIR = ROOT / "output"
OUTPUT_XLSX = OUTPUT_DIR / "звонки_менеджеры_март-05июня_анализ.xlsx"

POSITIVE_STAGES = {
    "Оплачен",
    "Запущено в производство",
    "Счет отдан в оплату",
    "Передано в логистику",
    "Передано дизайнеру",
    "УПД передано на подписании",
    "Согласие получено",
    "КП получено клиентом",
    "Потребность выявлена",
    "Назначен Ответственный",
}

NEGATIVE_STAGES = {
    "Отказ ms",
    "Проиграли по цене",
    "Низкий бюджет",
    "Недозвон",
    "Отложено",
    "Не удалось обработать",
    "Отмена мероприятия",
    "Отсутствие продукции ms",
    "Просрочено",
}

STAGE_WEIGHTS = {
    "positive": 0.2,
    "neutral": 0.6,
    "negative": 1.0,
}

ISSUES = [
    (
        "price_objection",
        "Цена / бюджет",
        [
            "дорог",
            "цена",
            "стоим",
            "бюджет",
            "дешев",
            "скид",
            "прайс",
        ],
        "Рано переводить цену в ценность: показывать пакеты, якорить вариантами и не давать скидку без обмена на условия.",
    ),
    (
        "postpone",
        "Перенос / подумать",
        [
            "подумаю",
            "перезвон",
            "созвоним",
            "не актуаль",
            "отлож",
        ],
        "Фиксировать следующий контакт с датой и временем, иначе сделка зависает без ответственности.",
    ),
    (
        "compare",
        "Сравнение с конкурентами",
        [
            "конкурент",
            "сравн",
            "аналог",
            "другой вариант",
            "вариантов",
            "где дешевле",
        ],
        "Отстраивать ценность по срокам, комплектации, сервису и примерам, а не спорить только по цене.",
    ),
    (
        "qualification_gap",
        "Недоквалификация",
        [
            "размер",
            "тираж",
            "срок",
            "доставка",
            "монтаж",
            "адрес",
            "материал",
            "оплата",
            "формат",
            "бюджет",
            "кто принимает решение",
            "лпр",
        ],
        "Задавать обязательный чек-лист вопросов по задаче, срокам, ЛПР, доставке и бюджету до отправки КП.",
    ),
    (
        "no_next_step",
        "Нет следующего шага",
        [
            "отправлю",
            "скину",
            "пришлю",
            "согласуем",
            "оформим",
            "подберу",
            "выберем",
            "жду ответ",
            "до связи",
            "созвон",
        ],
        "Закрывать каждый звонок на конкретное действие менеджера и клиента, иначе теряется контроль над сделкой.",
    ),
    (
        "upsell_missed",
        "Упущенный апсейл",
        [
            "монтаж",
            "дизайн",
            "доставка",
            "сроч",
            "сборк",
            "установк",
            "макет",
        ],
        "Пакетировать допуслуги и предлагать комплекс: дизайн, доставка, монтаж, срочное производство.",
    ),
]

ISSUE_KIND = {
    "price_objection": "problem",
    "postpone": "problem",
    "compare": "problem",
    "qualification_gap": "problem",
    "no_next_step": "problem",
    "upsell_missed": "growth",
}

CORE_PRODUCT_PATTERNS = {
    "роллап": ["роллап", "ролл-ап", "ролик"],
    "стенд": ["стенд", "стенда", "стенды"],
    "пресс-волл": ["пресс-волл", "press wall"],
    "баннер": ["баннер"],
    "флаг": ["флаг", "виндер"],
    "лайтбокс": ["лайтбокс", "световой бокс", "световой короб"],
    "буквы": ["букв", "объемн"],
    "вывеска": ["вывеск"],
    "табличка": ["табличк"],
    "печать": ["печать", "полиграф", "листовк", "плакат", "наклейк"],
    "стойка": ["стойк", "промостойк", "поп-ап"],
    "павильон": ["павильон", "застройк"],
}

SERVICE_PATTERNS = [
    "доставк",
    "монтаж",
    "установк",
    "сборк",
    "дизайн",
    "макет",
    "срочн",
]

QUALIFICATION_FIELDS = [
    ("срок", ["срок", "дата", "когда", "до пятниц", "до понедель", "к концу", "сегодня", "завтра"]),
    ("тираж", ["тираж", "количеств", "штук"]),
    ("размер", ["размер", "см", "метр", "х"]),
    ("доставка", ["доставк", "привез", "забер"]),
    ("монтаж", ["монтаж", "демонтаж", "установк", "сборк"]),
    ("адрес", ["адрес", "город", "куда"]),
    ("материал", ["материал", "пвх", "композит", "акрил", "алюмин"]),
    ("оплата", ["оплат", "счет", "предоплат"]),
    ("бюджет", ["бюджет", "стоим", "цена", "дорог"]),
    ("ЛПР", ["кто принимает решение", "руководител", "директор", "согласован", "лпр"]),
]

CLIENT_SIGNAL_RULES = [
    ("Цена / бюджет", ["дорог", "цена", "стоим", "бюджет", "скид", "прайс"], "клиент давит по цене или бюджету"),
    ("Перенос / подумать", ["подумаю", "перезвон", "созвоним", "не актуаль", "отлож"], "решение откладывается"),
    ("Сравнение с конкурентами", ["конкурент", "сравн", "аналог", "дешевле"], "клиент сравнивает с альтернативами"),
    ("Недоквалификация", ["размер", "тираж", "срок", "доставка", "монтаж", "адрес", "материал", "оплата", "бюджет"], "не хватает вводных по задаче"),
    ("Упущенный апсейл", ["монтаж", "дизайн", "доставка", "сроч", "сборк", "установк", "макет"], "есть запрос на сервис или комплекс, но он не монетизирован"),
]

MANAGER_ACTION_RULES = {
    "price_objection": "не перевел цену в ценность",
    "postpone": "не закрепил дату следующего шага",
    "compare": "не отстроился от альтернатив по срокам и сервису",
    "qualification_gap": "не добрал ключевые вводные по ТЗ",
    "no_next_step": "не закрыл звонок на конкретный next step",
    "upsell_missed": "не предложил пакет услуг и допродажу",
}


def normalize(text: str) -> str:
    text = text.lower().replace("ё", "е")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def stage_class(stage: str) -> str:
    if stage in POSITIVE_STAGES:
        return "positive"
    if stage in NEGATIVE_STAGES:
        return "negative"
    return "neutral"


def stage_weight(stage: str) -> float:
    return STAGE_WEIGHTS[stage_class(stage)]


def read_json_payload(json_path: Path) -> dict:
    with json_path.open("r", encoding="utf-8") as fh:
        return json.load(fh)


def payload_text(payload: dict) -> str:
    return normalize(" ".join(segment.get("text", "") for segment in payload.get("segments", [])))


def first_hit(text: str, patterns: Iterable[str]) -> str:
    for pattern in patterns:
        if pattern in text:
            return pattern
    return ""


def extract_product_families(text: str) -> list[str]:
    result: list[str] = []
    for label, patterns in CORE_PRODUCT_PATTERNS.items():
        if first_hit(text, patterns):
            result.append(label)
    return result


def has_any(text: str, patterns: Iterable[str]) -> bool:
    return any(pattern in text for pattern in patterns)


def unique_join(values: Iterable[str]) -> str:
    seen: set[str] = set()
    ordered: list[str] = []
    for value in values:
        if value and value not in seen:
            seen.add(value)
            ordered.append(value)
    return "; ".join(ordered)


def qualification_missing_fields(text: str) -> list[str]:
    missing: list[str] = []
    for field_name, patterns in QUALIFICATION_FIELDS:
        if not has_any(text, patterns):
            missing.append(field_name)
    return missing


def summarize_client_signals(text: str, issues: list[dict[str, str | float]]) -> str:
    labels: list[str] = []
    issue_names = {str(issue["issue_name"]) for issue in issues}
    for label, patterns, description in CLIENT_SIGNAL_RULES:
        if label in issue_names and has_any(text, patterns):
            labels.append(description)
    if has_any(text, ["срочно", "сегодня", "завтра", "до пятниц", "до понедел", "к концу"]):
        labels.append("есть срочность по срокам")
    return unique_join(labels) or "явных клиентских возражений не зафиксировано"


def summarize_manager_actions(text: str, issues: list[dict[str, str | float]]) -> str:
    actions: list[str] = []
    issue_ids = [str(issue["issue_id"]) for issue in issues]
    for issue_id in issue_ids:
        action = MANAGER_ACTION_RULES.get(issue_id)
        if issue_id == "qualification_gap":
            missing = qualification_missing_fields(text)
            if missing:
                action = f"не добрал вводные по {', '.join(missing[:3])}"
        if action:
            actions.append(action)
    if not actions:
        actions.append("диалог без явного провала в скрипте")
    return unique_join(actions[:3])


def extract_excerpt(segments: list[dict[str, object]], issue_patterns: list[str]) -> str:
    normalized_segments: list[str] = []
    for segment in segments:
        text = str(segment.get("text", "")).strip()
        if text:
            normalized_segments.append(text)
    if not normalized_segments:
        return ""

    for text in normalized_segments:
        normalized = normalize(text)
        if any(pattern in normalized for pattern in issue_patterns):
            return text[:220]

    excerpt = " / ".join(normalized_segments[:2])
    return excerpt[:220]


def detect_issues(text: str, stage: str, product_families: list[str]) -> list[dict[str, str | float]]:
    issues: list[dict[str, str | float]] = []
    family_count = len(product_families)
    service_hit = has_any(text, SERVICE_PATTERNS)

    for issue_id, issue_name, patterns, recommendation in ISSUES:
        if issue_id == "qualification_gap":
            qualification_tokens = [
                "размер",
                "тираж",
                "срок",
                "доставка",
                "монтаж",
                "адрес",
                "материал",
                "оплата",
                "бюджет",
                "лпр",
                "кто принимает решение",
            ]
            qual_score = sum(1 for token in qualification_tokens if token in text)
            if qual_score < 1:
                issues.append(
                    {
                        "issue_id": issue_id,
                        "issue_name": issue_name,
                        "trigger": "мало квалифицирующих вопросов",
                        "manager_action": "не добрался до обязательной квалификации",
                        "risk": stage_weight(stage),
                        "recommendation": recommendation,
                        "kind": ISSUE_KIND[issue_id],
                    }
                )
            continue

        if issue_id == "no_next_step":
            if stage_class(stage) != "positive" and not has_any(text, patterns):
                issues.append(
                    {
                        "issue_id": issue_id,
                        "issue_name": issue_name,
                        "trigger": "нет зафиксированного next step",
                        "manager_action": "не закрыл звонок на следующий шаг",
                        "risk": stage_weight(stage),
                        "recommendation": recommendation,
                        "kind": ISSUE_KIND[issue_id],
                    }
                )
            continue

        if issue_id == "upsell_missed":
            if family_count >= 1 and not service_hit:
                issues.append(
                    {
                        "issue_id": issue_id,
                        "issue_name": issue_name,
                        "trigger": "одиночная продажа без допуслуг",
                        "manager_action": "не предложил пакетирование и допродажу",
                        "risk": 0.5 if stage_class(stage) == "positive" else stage_weight(stage),
                        "recommendation": recommendation,
                        "kind": ISSUE_KIND[issue_id],
                    }
                )
            continue

        if has_any(text, patterns):
            trigger = first_hit(text, patterns)
            manager_action = {
                "price_objection": "не перевел разговор от цены к ценности",
                "postpone": "не дожал сделку до конкретной даты",
                "compare": "не отстроился от конкурентов",
            }[issue_id]
            risk = stage_weight(stage)
            if issue_id == "price_objection" and stage in {"Проиграли по цене", "Низкий бюджет"}:
                risk *= 1.4
            if issue_id == "postpone" and stage in {"Отложено", "Просрочено"}:
                risk *= 1.3
            if issue_id == "compare" and stage in {"Проиграли по цене", "Отказ ms"}:
                risk *= 1.2
            issues.append(
                {
                    "issue_id": issue_id,
                    "issue_name": issue_name,
                    "trigger": trigger,
                    "manager_action": manager_action,
                    "risk": risk,
                    "recommendation": recommendation,
                    "kind": ISSUE_KIND[issue_id],
                }
            )

    issues.sort(key=lambda item: float(item["risk"]), reverse=True)
    return issues


@dataclass
class CallRecord:
    row_number: int
    file_stem: str
    manager_name: str
    direction: str
    duration_seconds: int
    date_text: str
    month_key: str
    stage: str
    stage_group: str
    text: str
    issues: list[dict[str, str | float]]
    product_families: list[str]
    client_signals: str
    manager_actions: str
    dialogue_excerpt: str


def parse_month(date_text: str) -> str:
    try:
        dt = datetime.strptime(date_text, "%d.%m.%Y %H:%M:%S")
    except ValueError:
        return date_text[:7] if len(date_text) >= 7 else date_text
    return dt.strftime("%m.%Y")


def build_records(ws) -> list[CallRecord]:
    records: list[CallRecord] = []
    for row_idx in range(2, ws.max_row + 1):
        filename = ws[f"R{row_idx}"].value or ""
        file_stem = Path(str(filename)).name.replace(".mp3", "")
        json_path = INPUT_JSON_DIR / f"{file_stem}.json"
        payload = read_json_payload(json_path) if json_path.exists() else {"segments": []}
        text = payload_text(payload)
        segments = payload.get("segments", []) if isinstance(payload, dict) else []
        stage = str(ws[f"L{row_idx}"].value or "")
        product_families = extract_product_families((ws[f"K{row_idx}"].value or "") + " " + text)
        issues = detect_issues(text, stage, product_families)
        client_signals = summarize_client_signals(text, issues)
        manager_actions = summarize_manager_actions(text, issues)
        issue_patterns = []
        for issue in issues:
            issue_id = str(issue["issue_id"])
            for stored_id, _, patterns, _ in ISSUES:
                if stored_id == issue_id:
                    issue_patterns.extend(patterns)
                    break
        dialogue_excerpt = extract_excerpt(segments, issue_patterns)
        duration_raw = ws[f"E{row_idx}"].value or 0
        try:
            duration_seconds = int(duration_raw)
        except (TypeError, ValueError):
            duration_seconds = 0
        records.append(
            CallRecord(
                row_number=row_idx,
                file_stem=file_stem,
                manager_name=str(ws[f"B{row_idx}"].value or ""),
                direction=str(ws[f"D{row_idx}"].value or ""),
                duration_seconds=duration_seconds,
                date_text=str(ws[f"A{row_idx}"].value or ""),
                month_key=parse_month(str(ws[f"A{row_idx}"].value or "")),
                stage=stage,
                stage_group=stage_class(stage),
                text=text,
                issues=issues,
                product_families=product_families,
                client_signals=client_signals,
                manager_actions=manager_actions,
                dialogue_excerpt=dialogue_excerpt,
            )
        )
    return records


def enrich_calls_sheet(ws, records: list[CallRecord]) -> list[str]:
    new_headers = [
        "Ключевая проблема",
        "Тип сигнала",
        "Сигналы клиента",
        "Возражение / триггер",
        "Что сделал / не сделал менеджер",
        "Фрагмент диалога",
        "Риск",
        "Рекомендация",
        "Потенциал апсейла",
        "Категории сделки",
    ]
    start_col = ws.max_column + 1
    for offset, header in enumerate(new_headers, start=start_col):
        ws.cell(row=1, column=offset, value=header)

    for record in records:
        primary = record.issues[0] if record.issues else None
        issue_names = "; ".join(issue["issue_name"] for issue in record.issues[:3]) or "Нет критичных проблем"
        signal_type = issue_kind_label(str(primary["issue_id"])) if primary else "Нет сигнала"
        client_signals = record.client_signals
        trigger = primary["trigger"] if primary else "Нет явного триггера"
        action = record.manager_actions if record.manager_actions else "Сделка движется без выраженного узкого места"
        excerpt = record.dialogue_excerpt or "Фрагмент не распознан"
        risk = round(float(primary["risk"]), 2) if primary else 0.0
        recommendation = primary["recommendation"] if primary else ""
        upsell = "Да" if "Упущенный апсейл" in issue_names else "Нет"
        categories = ", ".join(record.product_families) if record.product_families else "не распознано"
        values = [issue_names, signal_type, client_signals, trigger, action, excerpt, risk, recommendation, upsell, categories]
        for idx, value in enumerate(values, start=start_col):
            ws.cell(row=record.row_number, column=idx, value=value)

    return new_headers


def style_calls_sheet(ws, total_cols: int) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=total_cols).coordinate}"
    widths = {
        "A": 18,
        "B": 20,
        "C": 12,
        "D": 14,
        "E": 16,
        "F": 14,
        "G": 14,
        "H": 38,
        "I": 12,
        "J": 12,
        "K": 28,
        "L": 24,
        "M": 14,
        "N": 26,
        "O": 34,
        "P": 34,
        "Q": 16,
        "R": 42,
        "S": 30,
        "T": 34,
        "U": 34,
        "V": 42,
        "W": 26,
        "X": 16,
        "Y": 60,
        "Z": 12,
        "AA": 28,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width
    for col_idx in range(19, total_cols + 1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        if letter not in widths:
            ws.column_dimensions[letter].width = 24
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill("solid", fgColor="D9E2F3")


def aggregate(records: list[CallRecord]):
    issue_stats = defaultdict(lambda: Counter())
    monthly_stats = defaultdict(lambda: Counter())
    manager_stats = defaultdict(lambda: Counter())
    stage_stats = Counter()

    for record in records:
        month = record.month_key
        stage_group = record.stage_group
        monthly_stats[month]["calls"] += 1
        monthly_stats[month][f"{stage_group}_calls"] += 1
        stage_stats[stage_group] += 1

        has_problem_issue = False
        has_growth_issue = False
        has_any_issue = False
        for issue in record.issues:
            if str(issue.get("kind")) == "problem":
                has_problem_issue = True
            else:
                has_growth_issue = True
            has_any_issue = True

        manager = record.manager_name or "Не указан"
        manager_stats[manager]["calls"] += 1
        manager_stats[manager][f"{stage_group}_calls"] += 1
        manager_stats[manager]["duration_seconds"] += record.duration_seconds
        if has_problem_issue:
            manager_stats[manager]["problem_calls"] += 1
        if has_growth_issue:
            manager_stats[manager]["growth_calls"] += 1
        if has_any_issue:
            manager_stats[manager]["issue_calls"] += 1
        monthly_stats[month]["problem_calls"] += int(has_problem_issue)
        monthly_stats[month]["growth_calls"] += int(has_growth_issue)
        monthly_stats[month]["issue_calls"] += int(has_any_issue)

        seen = set()
        for issue in record.issues:
            issue_id = str(issue["issue_id"])
            if issue_id in seen:
                continue
            seen.add(issue_id)
            issue_stats[issue_id]["frequency"] += 1
            issue_stats[issue_id][f"{stage_group}_calls"] += 1
            issue_stats[issue_id]["risk_score"] += float(issue["risk"])
            monthly_stats[month][issue_id] += 1

    return issue_stats, monthly_stats, manager_stats, stage_stats


def issue_display_name(issue_id: str) -> str:
    for stored_id, issue_name, _, _ in ISSUES:
        if stored_id == issue_id:
            return issue_name
    return issue_id


def issue_recommendation(issue_id: str) -> str:
    for stored_id, _, _, recommendation in ISSUES:
        if stored_id == issue_id:
            return recommendation
    return ""


def issue_kind_label(issue_id: str) -> str:
    return "Проблема" if ISSUE_KIND.get(issue_id) == "problem" else "Точка роста"


def build_summary_sheet(ws, records: list[CallRecord], issue_stats, monthly_stats, manager_stats, stage_stats) -> None:
    ws.delete_rows(1, ws.max_row)
    ws["A1"] = "Параметр"
    ws["B1"] = "Значение"

    total_calls = len(records)
    problem_calls = sum(1 for r in records if any(str(issue.get("kind")) == "problem" for issue in r.issues))
    growth_calls = sum(1 for r in records if any(str(issue.get("kind")) == "growth" for issue in r.issues))
    any_signal_calls = sum(1 for r in records if r.issues)
    no_signal_calls = total_calls - any_signal_calls
    positive_calls = sum(1 for r in records if r.stage_group == "positive")
    negative_calls = sum(1 for r in records if r.stage_group == "negative")
    neutral_calls = sum(1 for r in records if r.stage_group == "neutral")
    avg_signal_count = sum(len(r.issues) for r in records) / total_calls if total_calls else 0
    avg_duration_sec = sum(r.duration_seconds for r in records) / total_calls if total_calls else 0

    meta = [
        ("Период", "01.03.2026 — 05.06.2026"),
        ("Мин. длительность", ">60 сек"),
        ("Только с записью", "Да"),
        ("Всего звонков", total_calls),
        ("Проблемных звонков", problem_calls),
        ("Звонков с точками роста", growth_calls),
        ("Звонков без сигналов", no_signal_calls),
        ("Позитивные звонки", positive_calls),
        ("Нейтральные звонки", neutral_calls),
        ("Негативные звонки", negative_calls),
        ("Доля позитивных звонков", f"{positive_calls / total_calls:.1%}"),
        ("Доля звонков без сигналов", f"{no_signal_calls / total_calls:.1%}"),
        ("Среднее сигналов на звонок", f"{avg_signal_count:.2f}"),
        ("Средняя длительность", f"{round(avg_duration_sec / 60, 1)} мин"),
        ("Доля комплексных сделок", f"{sum(1 for r in records if len(r.product_families) >= 2) / total_calls:.1%}"),
        ("Методика подсчета", "Проблемы и точки роста считаются отдельно; один звонок может попадать в несколько сигналов."),
    ]
    row = 2
    for key, value in meta:
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=value)
        row += 1

    start = row + 2
    headers = [
        "Ранг",
        "Проблема",
        "Тип",
        "Частота",
        "Доля всех звонков",
        "Доля проблемных звонков",
        "Средний риск",
        "Решение",
        "Позитив",
        "Нейтрал",
        "Негатив",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=start, column=col, value=header)

    ranked = sorted(issue_stats.items(), key=lambda item: (item[1]["risk_score"], item[1]["frequency"]), reverse=True)
    summary_rows = []
    for rank, (issue_id, stats) in enumerate(ranked, start=1):
        freq = stats["frequency"]
        avg_risk = stats["risk_score"] / freq if freq else 0
        kind_label = issue_kind_label(issue_id)
        if kind_label == "Проблема":
            share_label = f"{freq / problem_calls:.1%}" if problem_calls else "0.0%"
        else:
            share_label = f"{freq / growth_calls:.1%}" if growth_calls else "0.0%"
        summary_rows.append(
            [
                rank,
                issue_display_name(issue_id),
                kind_label,
                freq,
                f"{freq / total_calls:.1%}",
                share_label,
                round(avg_risk, 2),
                issue_recommendation(issue_id),
                stats["positive_calls"],
                stats["neutral_calls"],
                stats["negative_calls"],
            ]
        )

    for r_offset, values in enumerate(summary_rows, start=start + 1):
        for c_offset, value in enumerate(values, start=1):
            ws.cell(row=r_offset, column=c_offset, value=value)

    for cell in ws[start]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAD3")

    chart1 = BarChart()
    chart1.type = "bar"
    chart1.style = 10
    chart1.title = "Ключевые проблемы по частоте"
    chart1.y_axis.title = "Проблема"
    chart1.x_axis.title = "Частота"
    data = Reference(ws, min_col=3, min_row=start, max_row=start + min(8, len(summary_rows)))
    cats = Reference(ws, min_col=2, min_row=start + 1, max_row=start + min(8, len(summary_rows)))
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.height = 8
    chart1.width = 14
    ws.add_chart(chart1, "K2")

    stage_start = start + len(summary_rows) + 3
    ws.cell(row=stage_start, column=1, value="Статус")
    ws.cell(row=stage_start, column=2, value="Звонки")
    ws.cell(row=stage_start, column=3, value="Доля")
    stage_rows = [
        ("Позитивные", stage_stats["positive"], f"{stage_stats['positive'] / total_calls:.1%}"),
        ("Нейтральные", stage_stats["neutral"], f"{stage_stats['neutral'] / total_calls:.1%}"),
        ("Негативные", stage_stats["negative"], f"{stage_stats['negative'] / total_calls:.1%}"),
    ]
    for idx, values in enumerate(stage_rows, start=stage_start + 1):
        for col, value in enumerate(values, start=1):
            ws.cell(row=idx, column=col, value=value)

    chart2 = BarChart()
    chart2.type = "bar"
    chart2.style = 11
    chart2.title = "Исходы звонков"
    chart2.y_axis.title = "Статус"
    chart2.x_axis.title = "Звонки"
    data = Reference(ws, min_col=2, min_row=stage_start, max_row=stage_start + len(stage_rows))
    cats = Reference(ws, min_col=1, min_row=stage_start + 1, max_row=stage_start + len(stage_rows))
    chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(cats)
    chart2.height = 6
    chart2.width = 12
    ws.add_chart(chart2, "K18")
    for cell in ws[stage_start]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAD3")

    month_order = sorted(monthly_stats.keys(), key=lambda m: datetime.strptime("01." + m, "%d.%m.%Y"))
    trend_start = stage_start + len(stage_rows) + 4
    ws.cell(row=trend_start, column=1, value="Месяц")
    ws.cell(row=trend_start, column=2, value="Проблемные")
    ws.cell(row=trend_start, column=3, value="Точки роста")
    ws.cell(row=trend_start, column=4, value="Позитивные")
    ws.cell(row=trend_start, column=5, value="Всего")
    ws.cell(row=trend_start, column=6, value="Доля проблемных")

    monthly_rows = []
    for month in month_order:
        calls = monthly_stats[month]["calls"]
        problem = monthly_stats[month]["problem_calls"]
        growth = monthly_stats[month]["growth_calls"]
        positive = monthly_stats[month]["positive_calls"]
        monthly_rows.append((month, problem, growth, positive, calls, f"{problem / calls:.1%}" if calls else "0.0%"))

    for idx, values in enumerate(monthly_rows, start=trend_start + 1):
        for col, value in enumerate(values, start=1):
            ws.cell(row=idx, column=col, value=value)

    chart3 = LineChart()
    chart3.title = "Тренд проблемных и позитивных звонков"
    chart3.y_axis.title = "Звонки"
    chart3.x_axis.title = "Месяц"
    data = Reference(ws, min_col=2, min_row=trend_start, max_col=4, max_row=trend_start + len(monthly_rows))
    cats = Reference(ws, min_col=1, min_row=trend_start + 1, max_row=trend_start + len(monthly_rows))
    chart3.add_data(data, titles_from_data=True)
    chart3.set_categories(cats)
    chart3.height = 7
    chart3.width = 14
    if chart3.series:
        chart3.series[0].trendline = Trendline(trendlineType="linear")
    ws.add_chart(chart3, "K32")
    for cell in ws[trend_start]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FCE5CD")

    manager_start = trend_start + len(monthly_rows) + 4
    ws.cell(row=manager_start, column=1, value="Менеджер")
    ws.cell(row=manager_start, column=2, value="Звонки")
    ws.cell(row=manager_start, column=3, value="Проблемные")
    ws.cell(row=manager_start, column=4, value="Точки роста")
    ws.cell(row=manager_start, column=5, value="Доля проблемных")
    ws.cell(row=manager_start, column=6, value="Позитивные")
    ws.cell(row=manager_start, column=7, value="Негативные")
    ws.cell(row=manager_start, column=8, value="Средняя длительность")

    manager_rows = []
    ranked_managers = sorted(
        manager_stats.items(),
        key=lambda item: (item[1]["problem_calls"] / item[1]["calls"] if item[1]["calls"] else 0, item[1]["problem_calls"]),
        reverse=True,
    )
    for manager, stats in ranked_managers:
        calls = stats["calls"]
        if calls < 10:
            continue
        avg_minutes = stats["duration_seconds"] / calls / 60 if calls else 0
        manager_rows.append(
            (
                manager,
                calls,
                stats["problem_calls"],
                stats["growth_calls"],
                f"{stats['problem_calls'] / calls:.1%}" if calls else "0.0%",
                stats["positive_calls"],
                stats["negative_calls"],
                f"{avg_minutes:.1f} мин",
            )
        )
    if not manager_rows:
        for manager, stats in sorted(manager_stats.items(), key=lambda item: item[1]["calls"], reverse=True)[:8]:
            calls = stats["calls"]
            avg_minutes = stats["duration_seconds"] / calls / 60 if calls else 0
            manager_rows.append(
                (
                    manager,
                    calls,
                    stats["problem_calls"],
                    stats["growth_calls"],
                    f"{stats['problem_calls'] / calls:.1%}" if calls else "0.0%",
                    stats["positive_calls"],
                    stats["negative_calls"],
                    f"{avg_minutes:.1f} мин",
                )
            )
    manager_rows = manager_rows[:8]
    for idx, values in enumerate(manager_rows, start=manager_start + 1):
        for col, value in enumerate(values, start=1):
            ws.cell(row=idx, column=col, value=value)

    chart4 = BarChart()
    chart4.type = "bar"
    chart4.style = 12
    chart4.title = "Топ менеджеров по доле проблемных звонков"
    chart4.y_axis.title = "Менеджер"
    chart4.x_axis.title = "Доля проблемных"
    data = Reference(ws, min_col=5, min_row=manager_start, max_row=manager_start + len(manager_rows))
    cats = Reference(ws, min_col=1, min_row=manager_start + 1, max_row=manager_start + len(manager_rows))
    chart4.add_data(data, titles_from_data=True)
    chart4.set_categories(cats)
    chart4.height = 7
    chart4.width = 16
    ws.add_chart(chart4, "K60")
    for cell in ws[manager_start]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAD3")

    product_counter = Counter()
    for record in records:
        for family in record.product_families:
            product_counter[family] += 1
    product_start = manager_start + len(manager_rows) + 4
    ws.cell(row=product_start, column=1, value="Категория сделки")
    ws.cell(row=product_start, column=2, value="Звонки")
    ws.cell(row=product_start, column=3, value="Доля")
    product_rows = []
    for family, count in product_counter.most_common(6):
        product_rows.append((family, count, f"{count / total_calls:.1%}" if total_calls else "0.0%"))
    for idx, values in enumerate(product_rows, start=product_start + 1):
        for col, value in enumerate(values, start=1):
            ws.cell(row=idx, column=col, value=value)
    for cell in ws[product_start]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FCE5CD")

    chart5 = BarChart()
    chart5.type = "bar"
    chart5.style = 13
    chart5.title = "Топ продуктовых категорий"
    chart5.y_axis.title = "Категория"
    chart5.x_axis.title = "Звонки"
    data = Reference(ws, min_col=2, min_row=product_start, max_row=product_start + len(product_rows))
    cats = Reference(ws, min_col=1, min_row=product_start + 1, max_row=product_start + len(product_rows))
    chart5.add_data(data, titles_from_data=True)
    chart5.set_categories(cats)
    chart5.height = 7
    chart5.width = 16
    ws.add_chart(chart5, "K78")

    for col in range(1, 10):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 24
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["F"].width = 60
    ws.freeze_panes = "A2"
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            if cell.value is not None and cell.row != 1:
                cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_forecast_sheet(ws, records: list[CallRecord], issue_stats) -> None:
    ws.delete_rows(1, ws.max_row)
    ws["A1"] = "Метрика"
    ws["B1"] = "Текущее"
    ws["C1"] = "Консервативно"
    ws["D1"] = "Базово"
    ws["E1"] = "Агрессивно"

    total = len(records)
    positive = sum(1 for r in records if r.stage_group == "positive")
    complex_calls = sum(1 for r in records if len(r.product_families) >= 2)
    single_core_calls = sum(1 for r in records if len(r.product_families) == 1)

    current_conv = positive / total
    current_complex = complex_calls / total
    current_single = single_core_calls / total

    price_share = issue_stats["price_objection"]["frequency"] / total if issue_stats["price_objection"]["frequency"] else 0
    postpone_share = issue_stats["postpone"]["frequency"] / total if issue_stats["postpone"]["frequency"] else 0
    qual_share = issue_stats["qualification_gap"]["frequency"] / total if issue_stats["qualification_gap"]["frequency"] else 0
    upsell_share = issue_stats["upsell_missed"]["frequency"] / total if issue_stats["upsell_missed"]["frequency"] else 0

    rows = [
        ("Конверсия звонок→сделка", current_conv, current_conv + min(0.035, price_share * 0.20 * 0.25 + postpone_share * 0.15 * 0.20), current_conv + min(0.060, price_share * 0.30 * 0.35 + postpone_share * 0.20 * 0.25 + qual_share * 0.15 * 0.15), current_conv + min(0.100, price_share * 0.40 * 0.45 + postpone_share * 0.25 * 0.30 + qual_share * 0.20 * 0.25)),
        ("Доля комплексных сделок", current_complex, current_complex + min(0.050, current_single * 0.25), current_complex + min(0.080, current_single * 0.35), current_complex + min(0.120, current_single * 0.45)),
        ("Доля звонков с апсейлом", upsell_share, upsell_share + 0.03, upsell_share + 0.05, upsell_share + 0.08),
    ]

    for row_idx, (metric, current, conservative, base, aggressive) in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=metric)
        ws.cell(row=row_idx, column=2, value=round(current, 1) if current > 1 else f"{current:.1%}")
        ws.cell(row=row_idx, column=3, value=f"{conservative:.1%}")
        ws.cell(row=row_idx, column=4, value=f"{base:.1%}")
        ws.cell(row=row_idx, column=5, value=f"{aggressive:.1%}")

    assumptions_row = 7
    ws.cell(row=assumptions_row, column=1, value="Допущения")
    ws.cell(row=assumptions_row + 1, column=1, value="Снижение price objection")
    ws.cell(row=assumptions_row + 1, column=2, value="20% / 30% / 40%")
    ws.cell(row=assumptions_row + 2, column=1, value="Снижение postponed deals")
    ws.cell(row=assumptions_row + 2, column=2, value="15% / 20% / 25%")
    ws.cell(row=assumptions_row + 3, column=1, value="Рост complex-sale share")
    ws.cell(row=assumptions_row + 3, column=2, value="25% / 35% / 45% от single-core сделок")

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FCE5CD")
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(INPUT_XLSX)
    ws_calls = wb["Звонки"]
    ws_summary = wb["Сводка"]

    records = build_records(ws_calls)
    new_headers = enrich_calls_sheet(ws_calls, records)
    style_calls_sheet(ws_calls, ws_calls.max_column)

    issue_stats, monthly_stats, manager_stats, stage_stats = aggregate(records)

    build_summary_sheet(ws_summary, records, issue_stats, monthly_stats, manager_stats, stage_stats)
    if "Прогноз" in wb.sheetnames:
        del wb["Прогноз"]
    ws_forecast = wb.create_sheet("Прогноз")
    build_forecast_sheet(ws_forecast, records, issue_stats)

    wb.save(OUTPUT_XLSX)
    print(f"saved {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
